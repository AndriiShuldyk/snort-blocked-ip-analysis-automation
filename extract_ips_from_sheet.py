import os
import re
import datetime
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options

# Configuration
OUTPUT_DIR = r""
MASTER_XLSX_PATH = os.path.join(OUTPUT_DIR, "master.xlsx")
RED_FILL_COLOR = "FFFF0000"

HIGHLIGHT_ORGS = ["Microsoft Corporation", "Google LLC", "Amazon.com", "Akamai"]

# Website credentials and configuration
WEBSITE_CREDENTIALS = {
    'username': '',
    'password': ''
}
BASE_URL = ""


def setup_chrome_driver():
    """
    Setup Chrome WebDriver with options to handle SSL and other configurations
    """
    # Chrome options
    chrome_options = Options()
    chrome_options.add_argument('--ignore-ssl-errors=yes')
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--disable-web-security')
    chrome_options.add_argument('--allow-running-insecure-content')
    
    # Reduce logging
    chrome_options.add_argument('--log-level=3')
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    
    # Try to use WebDriverManager if installed
    try:
        from webdriver_manager.chrome import ChromeDriverManager
        service = Service(ChromeDriverManager().install())
    except ImportError:
        # Fallback to manual path if WebDriverManager is not installed
        service = Service(r"C:\path\to\chromedriver.exe")  # Replace with your actual path
    
    # Create and return the WebDriver
    return webdriver.Chrome(service=service, options=chrome_options)


def login_to_website(driver, credentials):
    """
    Login to the website using provided credentials
    """
    try:
        # Navigate to the website
        print("Navigating to login page...")
        driver.get(BASE_URL)
        
        # Find and interact with username field
        username_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "usernamefld"))
        )
        username_field.clear()
        username_field.send_keys(credentials['username'])
        
        # Find and interact with password field
        password_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "passwordfld"))
        )
        password_field.clear()
        password_field.send_keys(credentials['password'])
        
        # Find and click login button
        login_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "input[name='login'][value='Sign In']"))
        )
        login_button.click()
        
        # Wait for login to complete
        print("Logging in...")
        time.sleep(3)
        
        return True
        
    except Exception as e:
        print(f"Login failed: {e}")
        return False


def add_ips_to_passlist(ip_list):
    """
    Add IP addresses to the Pass List using Selenium
    """
    if not ip_list:
        print("No IP addresses to add to Pass List.")
        return False
    
    print(f"Adding {len(ip_list)} IP addresses to Pass List...")
    
    # Initialize the WebDriver
    driver = setup_chrome_driver()
    
    try:
        # Login to the website
        if not login_to_website(driver, WEBSITE_CREDENTIALS):
            print("Failed to login to the website.")
            return False
        
        print("Successfully logged in to the website.")
        
        # Navigate to Pass List edit page
        passlist_url = "https://10.35.32.1:8443/snort/snort_passlist_edit.php?id=0"
        print(f"Navigating to Pass List page: {passlist_url}")
        driver.get(passlist_url)
        
        # Wait for page to load
        time.sleep(2)
        
        # Add each IP address
        for i, ip in enumerate(ip_list, 1):
            print(f"Adding IP {i}/{len(ip_list)}: {ip}")
            
            try:
                # Click "Add IP" button
                add_ip_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.ID, "addrow"))
                )
                add_ip_button.click()
                
                # Wait a bit for the new row to appear
                time.sleep(1)
                
                # Find all address input fields and get the last one (newest)
                address_fields = driver.find_elements(By.CSS_SELECTOR, "input[name^='address'][placeholder='Address']")
                
                if not address_fields:
                    print(f"No address fields found for IP: {ip}")
                    continue
                
                # Get the last (newest) address field
                address_field = address_fields[-1]
                
                # Get the field ID for debugging
                field_id = address_field.get_attribute('id')
                field_name = address_field.get_attribute('name')
                print(f"Using address field: ID={field_id}, Name={field_name}")
                
                # Clear the field and enter the IP address
                address_field.clear()
                address_field.send_keys(ip)
                
                print(f"Successfully added IP: {ip}")
                
                # Small delay between additions to avoid overwhelming the page
                time.sleep(0.5)
                
            except Exception as e:
                print(f"Error adding IP {ip}: {e}")
                # Try to take a screenshot for debugging
                try:
                    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                    screenshot_path = os.path.join(OUTPUT_DIR, f"error_screenshot_{timestamp}.png")
                    driver.save_screenshot(screenshot_path)
                    print(f"Error screenshot saved to: {screenshot_path}")
                except:
                    pass
                # Continue with next IP even if one fails
                continue
        
        # Click Save button
        print("Saving changes...")
        try:
            save_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, "save"))
            )
            save_button.click()
            
            # Wait for save to complete
            time.sleep(3)
            
            print("Successfully saved all IP addresses to Pass List!")
            return True
            
        except Exception as e:
            print(f"Error clicking Save button: {e}")
            return False
        
    except Exception as e:
        print(f"Error adding IPs to Pass List: {e}")
        return False
    
    finally:
        # Close the browser
        print("Closing browser...")
        if 'driver' in locals():
            driver.quit()


def extract_red_ips_from_newest_sheet(org_choice=None):
    """
    Extract IP addresses from cells with red highlight from the newest sheet
    in the Excel workbook and save them to a text file.
    
    Args:
        org_choice: Integer 1-5 for specific organization, 5 for all, None for prompt
    
    Returns:
        tuple: (bool: success, list: extracted IPs, str: organization name, str: output file path)
    """
    print("=" * 60)
    print("EXTRACTING RED-HIGHLIGHTED IP ADDRESSES")
    print("=" * 60)
    
    # Prompt for organization choice if not provided
    if org_choice is None:
        print("\nSelect organization to extract IPs for:")
        print("1 - Google LLC")
        print("2 - Microsoft Corporation")
        print("3 - Amazon.com")
        print("4 - Akamai")
        print("5 - All red-highlighted IPs")
        
        try:
            org_choice = int(input("\nEnter your choice (1-5): "))
            if org_choice < 1 or org_choice > 5:
                print("Invalid choice. Please enter a number between 1 and 5.")
                return False, [], "", ""
        except ValueError:
            print("Invalid input. Please enter a number between 1 and 5.")
            return False, [], "", ""
    
    # Map choice to organization name
    selected_org = None
    org_name_for_file = ""
    org_display_name = ""
    if org_choice == 1:
        selected_org = HIGHLIGHT_ORGS[1]  # Google LLC
        org_name_for_file = "google"
        org_display_name = "Google LLC"
    elif org_choice == 2:
        selected_org = HIGHLIGHT_ORGS[0]  # Microsoft Corporation
        org_name_for_file = "microsoft"
        org_display_name = "Microsoft Corporation"
    elif org_choice == 3:
        selected_org = HIGHLIGHT_ORGS[2]  # Amazon.com
        org_name_for_file = "amazon"
        org_display_name = "Amazon.com"
    elif org_choice == 4:
        selected_org = HIGHLIGHT_ORGS[3]  # Akamai
        org_name_for_file = "akamai"
        org_display_name = "Akamai"
    elif org_choice == 5:
        selected_org = None  # All organizations
        org_name_for_file = "all"
        org_display_name = "All red IPs"
    
    print(f"Selected organization: {org_display_name}")
    
    # Check if master file exists
    if not os.path.exists(MASTER_XLSX_PATH):
        print(f"Error: Master Excel file not found at {MASTER_XLSX_PATH}")
        return False, [], org_display_name, ""
    
    # Load the workbook
    print(f"Loading workbook: {MASTER_XLSX_PATH}")
    wb = load_workbook(MASTER_XLSX_PATH)
    
    # Get sheet names
    sheet_names = wb.sheetnames
    if not sheet_names:
        print("Error: No sheets found in the workbook")
        return False, [], org_display_name, ""
    
    # Find sheets with date pattern (dd_mm_yyyy)
    date_sheets = []
    for sheet_name in sheet_names:
        if re.match(r'^\d{2}_\d{2}_\d{4}$', sheet_name):
            # Convert sheet name to datetime for comparison
            try:
                sheet_date = datetime.datetime.strptime(sheet_name, '%d_%m_%Y')
                date_sheets.append((sheet_name, sheet_date))
            except ValueError:
                # Skip sheets that don't match our expected date format
                continue
    
    if not date_sheets:
        print("Error: No sheets with date pattern found")
        return False, [], org_display_name, ""
    
    # Sort sheets by date (newest first)
    date_sheets.sort(key=lambda x: x[1], reverse=True)
    newest_sheet_name = date_sheets[0][0]
    
    print(f"Found newest sheet: {newest_sheet_name} (Date: {date_sheets[0][1].strftime('%d-%m-%Y')})")
    
    # Get the newest sheet
    ws = wb[newest_sheet_name]
    
    # Find column indices for IP and organization
    ip_col_index = None
    org_col_index = None
    
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == 'ip':
            ip_col_index = idx
        elif cell.value == 'org':
            org_col_index = idx
    
    if not ip_col_index:
        print("Error: Could not find 'ip' column in the sheet")
        return False, [], org_display_name, ""
    
    if not org_col_index and selected_org:
        print("Warning: Could not find 'org' column in the sheet, but organization filtering was requested")
        print("Will fall back to checking all red rows")
        selected_org = None
    
    # Extract red-highlighted IP addresses
    red_ips = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Skip header row
        # Check if any cell in the row has red fill
        is_red = any(cell.fill.start_color.rgb == RED_FILL_COLOR for cell in row)
        
        if is_red:
            ip_cell = row[ip_col_index - 1]  # Adjusting for 0-based indexing
            
            # If we're filtering by organization, check the org field
            if selected_org and org_col_index:
                org_cell = row[org_col_index - 1]
                if org_cell.value and selected_org in org_cell.value:
                    if ip_cell.value:
                        red_ips.append(ip_cell.value)
            # Otherwise include all red IPs
            elif not selected_org:
                if ip_cell.value:
                    red_ips.append(ip_cell.value)
    
    print(f"Found {len(red_ips)} IP addresses matching your criteria")
    
    # Save IP addresses to text file with organization name in filename
    output_filename = f"red_ips_{newest_sheet_name}_{org_name_for_file}.txt"
    output_path = os.path.join(OUTPUT_DIR, output_filename)
    
    with open(output_path, 'w') as f:
        for ip in red_ips:
            f.write(f"{ip}\n")
    
    print(f"Saved IP addresses to: {output_path}")
    return True, red_ips, org_display_name, output_path


def delete_file_safely(file_path):
    """
    Safely delete a file if it exists
    """
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
            print(f"Deleted file: {file_path}")
            return True
        else:
            print(f"File not found for deletion: {file_path}")
            return False
    except Exception as e:
        print(f"Error deleting file {file_path}: {e}")
        return False


def main():
    """Main function to coordinate the workflow"""
    start_time = datetime.datetime.now()
    
    # Create output directory if it doesn't exist
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    success, extracted_ips, org_name, txt_file_path = extract_red_ips_from_newest_sheet()
    
    if success and extracted_ips:
        print("\n" + "=" * 60)
        print("IP EXTRACTION COMPLETED SUCCESSFULLY")
        print("=" * 60)
        
        # Ask user if they want to add IPs to Pass List
        print(f"\nExtracted {len(extracted_ips)} IP addresses for: {org_name}")
        print("IP addresses:")
        for i, ip in enumerate(extracted_ips[:10], 1):  # Show first 10 IPs
            print(f"  {i}. {ip}")
        if len(extracted_ips) > 10:
            print(f"  ... and {len(extracted_ips) - 10} more")
        
        while True:
            add_to_passlist = input(f"\nDo you want to add these {len(extracted_ips)} IP addresses to the Pass List? (y/n): ").strip().lower()
            if add_to_passlist in ['y', 'yes']:
                success_passlist = add_ips_to_passlist(extracted_ips)
                if success_passlist:
                    print("IPs successfully processed for Pass List addition.")
                    # Delete the txt file since IPs were added to pass list
                    print("Cleaning up temporary files...")
                    delete_file_safely(txt_file_path)
                else:
                    print("Failed to add IPs to Pass List.")
                    print(f"Text file with IPs preserved at: {txt_file_path}")
                break
            elif add_to_passlist in ['n', 'no']:
                print("Skipping Pass List addition.")
                print(f"Text file with IPs preserved at: {txt_file_path}")
                break
            else:
                print("Please enter 'y' for yes or 'n' for no.")
    
    elif success and not extracted_ips:
        print("\n" + "=" * 60)
        print("NO IP ADDRESSES FOUND")
        print("=" * 60)
        print("No red-highlighted IP addresses found matching your criteria.")
        # Delete empty txt file if it was created
        if txt_file_path and os.path.exists(txt_file_path):
            delete_file_safely(txt_file_path)
    else:
        print("\n" + "=" * 60)
        print("SCRIPT COMPLETED WITH ERRORS")
        print("=" * 60)
    
    elapsed_time = (datetime.datetime.now() - start_time).total_seconds()
    print(f"Total execution time: {elapsed_time:.2f} seconds")


if __name__ == "__main__":
    main()