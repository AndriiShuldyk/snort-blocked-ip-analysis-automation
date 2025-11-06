import os
import time
import csv
import glob
import tarfile
import re
import datetime
import ipaddress
import shutil
import pandas as pd
import ipinfo
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# ========== CONFIGURATION ==========
WEBSITE_CREDENTIALS = {
    'username': '',
    'password': ''
}

# ========== WEBSITE CONFIGURATION ==========
BASE_URL = ""
SNORT_BLOCKED_HOSTS_URL = BASE_URL + "snort/snort_blocked.php"

# ========== PATH CONFIGURATION ==========
DOWNLOAD_DIR = r""
OUTPUT_DIR = r""

# ========== IPINFO CONFIGURATION ==========
IPINFO_ACCESS_TOKEN = ''
IPINFO_FIELDS = ['ip', 'org', 'country_name', 'hostname']

# ========== HIGHLIGHTING CONFIGURATION ==========
HIGHLIGHT_ORGS = ["Microsoft Corporation", "Google LLC", "Amazon.com", "Akamai"]


def setup_chrome_driver():
    """
    Setup Chrome WebDriver with options to handle SSL and other configurations
    """
    # Create download directory if it doesn't exist
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    
    # Chrome options
    chrome_options = Options()
    chrome_options.add_argument('--ignore-ssl-errors=yes')
    chrome_options.add_argument('--ignore-certificate-errors')
    
    # Set up download preferences
    chrome_options.add_experimental_option("prefs", {
        "download.default_directory": DOWNLOAD_DIR,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": False
    })
    
    # Reduce logging
    chrome_options.add_argument('--log-level=3')
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    
    # Try to use WebDriverManager if installed
    try:
        from webdriver_manager.chrome import ChromeDriverManager
        service = Service(ChromeDriverManager().install())
    except ImportError:
        # Fallback to manual path if WebDriverManager is not installed
        service = Service(r"C:\path\to\chromedriver.exe")  # Replace with your actual path
    
    # Create and return the WebDriver
    return webdriver.Chrome(service=service, options=chrome_options)


def login_and_download_blocked_hosts(credentials):
    """
    Login to the website, navigate to Snort blocked hosts page, and download the file
    """
    # Validate credentials
    if not credentials['username'] or not credentials['password']:
        print("Error: Username or password not set")
        return None
    
    # Initialize the WebDriver
    driver = setup_chrome_driver()
    
    try:
        # Navigate to the website
        print("Navigating to login page...")
        driver.get(BASE_URL)
        
        # Find and interact with username field
        username_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "usernamefld"))
        )
        username_field.clear()
        username_field.send_keys(credentials['username'])
        
        # Find and interact with password field
        password_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "passwordfld"))
        )
        password_field.clear()
        password_field.send_keys(credentials['password'])
        
        # Find and click login button
        login_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "btn-success"))
        )
        login_button.click()
        
        # Wait for login to complete
        print("Logging in...")
        time.sleep(3)
        
        # Navigate to Snort blocked hosts page
        print("Navigating to Snort blocked hosts page...")
        driver.get(SNORT_BLOCKED_HOSTS_URL)
        
        # Wait for the page to load
        time.sleep(3)
        
        # Find and click the Download button
        download_button_locators = [
            (By.ID, "download"),
            (By.NAME, "download"),
            (By.XPATH, "//button[contains(text(), 'Download')]"),
            (By.XPATH, "//button[@title='Download interface log files as a gzip archive']"),
            (By.XPATH, "//button[contains(@class, 'btn-success') and @name='download']")
        ]
        
        download_button = None
        for locator in download_button_locators:
            try:
                download_button = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable(locator)
                )
                break
            except Exception:
                continue
        
        if download_button:
            print("Download button found. Initiating download...")
            download_button.click()
            
            # Wait for download to complete (adjust time as needed)
            print("Waiting for download to complete...")
            time.sleep(10)
            
            # Check if file exists in download directory
            files = os.listdir(DOWNLOAD_DIR)
            downloaded_files = [f for f in files if os.path.isfile(os.path.join(DOWNLOAD_DIR, f))]
            
            if downloaded_files:
                newest_file = max(
                    [os.path.join(DOWNLOAD_DIR, f) for f in downloaded_files],
                    key=os.path.getctime
                )
                print(f"Download completed: {os.path.basename(newest_file)}")
                print(f"File saved to: {newest_file}")
                return newest_file
            else:
                print("No downloaded files found")
                return None
        else:
            print("Download button not found")
            
            # Optional: Try to capture a screenshot to debug
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            screenshot_path = os.path.join(DOWNLOAD_DIR, f"debug_screenshot_{timestamp}.png")
            driver.save_screenshot(screenshot_path)
            print(f"Debug screenshot saved to: {screenshot_path}")
            
            return None
    
    except Exception as e:
        print(f"An error occurred: {e}")
        
        # Capture screenshot on error
        try:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            screenshot_path = os.path.join(DOWNLOAD_DIR, f"error_screenshot_{timestamp}.png")
            driver.save_screenshot(screenshot_path)
            print(f"Error screenshot saved to: {screenshot_path}")
        except:
            pass
            
        return None
    
    finally:
        # Close the browser
        print("Closing browser...")
        if 'driver' in locals():
            driver.quit()


def find_latest_download(downloads_folder, pattern="snort_blocked_*.tar.gz"):
    """Find the most recent downloaded snort file based on file modification time"""
    files = glob.glob(os.path.join(downloads_folder, pattern))
    if not files:
        # Try again with just any tar.gz file that might have "snort" in it
        files = glob.glob(os.path.join(downloads_folder, "*snort*.tar.gz"))
        
    if not files:
        # Last resort - try any tar.gz file
        files = glob.glob(os.path.join(downloads_folder, "*.tar.gz"))
        
    if not files:
        raise FileNotFoundError(f"No files matching {pattern} or similar found in {downloads_folder}")
    
    # Sort by modification time (newest first)
    latest_file = max(files, key=os.path.getmtime)
    print(f"Found latest file: {latest_file}")
    return latest_file


def find_previous_download(downloads_folder, latest_file, pattern="snort_blocked_*.tar.gz"):
    """Find the second most recent downloaded snort file based on file modification time"""
    files = glob.glob(os.path.join(downloads_folder, pattern))
    if not files:
        # Try again with just any tar.gz file that might have "snort" in it
        files = glob.glob(os.path.join(downloads_folder, "*snort*.tar.gz"))
        
    if not files:
        # Last resort - try any tar.gz file
        files = glob.glob(os.path.join(downloads_folder, "*.tar.gz"))
        
    if not files or len(files) < 2:
        print("No previous snort file found for comparison")
        return None
    
    # Filter out the latest file
    files = [f for f in files if f != latest_file]
    
    # Sort by modification time (newest first)
    previous_file = max(files, key=os.path.getmtime)
    print(f"Found previous file: {previous_file}")
    return previous_file


def cleanup_old_snort_files(downloads_folder, keep_latest=2, pattern="snort_blocked_*.tar.gz"):
    """Delete old snort files, keeping only the specified number of most recent files"""
    # Get all snort files
    files = glob.glob(os.path.join(downloads_folder, pattern))
    if not files:
        # Try again with just any tar.gz file that might have "snort" in it
        files = glob.glob(os.path.join(downloads_folder, "*snort*.tar.gz"))
    
    if not files:
        # Last resort - try any tar.gz file
        files = glob.glob(os.path.join(downloads_folder, "*.tar.gz"))
    
    if not files or len(files) <= keep_latest:
        print(f"No files to clean up (found {len(files)} files, keeping {keep_latest})")
        return
    
    # Sort files by modification time (newest first)
    sorted_files = sorted(files, key=os.path.getmtime, reverse=True)
    
    # Keep the newest 'keep_latest' files, delete the rest
    files_to_keep = sorted_files[:keep_latest]
    files_to_delete = sorted_files[keep_latest:]
    
    # Delete old files
    for file_path in files_to_delete:
        try:
            os.remove(file_path)
            print(f"Deleted old file: {os.path.basename(file_path)}")
        except Exception as e:
            print(f"Error deleting file {file_path}: {e}")
    
    print(f"Cleanup complete. Kept {len(files_to_keep)} recent files, deleted {len(files_to_delete)} old files.")


def extract_tar_gz(tar_gz_path, extract_dir):
    """Extract the tar.gz file to get the snort_block.pf file"""
    temp_dir = os.path.join(extract_dir, "temp_extract")
    
    # Create temp directory if it doesn't exist
    os.makedirs(temp_dir, exist_ok=True)
    
    print(f"Extracting {tar_gz_path}...")
    # List contents first to understand the file structure
    with tarfile.open(tar_gz_path, "r:gz") as tar:
        members = tar.getmembers()
        print(f"Archive contains {len(members)} files")
        
        # Debug: Print first few files to understand structure
        for i, member in enumerate(members[:5]):
            print(f"  - {member.name}")
            
        # Extract all files with filter='data' to address the deprecation warning
        tar.extractall(path=temp_dir, filter='data')
    
    # Look for either .tar files or directly for .pf files
    pf_files = glob.glob(os.path.join(temp_dir, "**", "*.pf"), recursive=True)
    tar_files = glob.glob(os.path.join(temp_dir, "**", "*.tar"), recursive=True)
    
    if pf_files:
        print(f"Found .pf file directly: {pf_files[0]}")
        pf_file = pf_files[0]
    elif tar_files:
        print(f"Found .tar file, extracting: {tar_files[0]}")
        with tarfile.open(tar_files[0], "r") as tar:
            tar.extractall(path=temp_dir, filter='data')
        # Look for .pf files again after extracting the .tar
        pf_files = glob.glob(os.path.join(temp_dir, "**", "*.pf"), recursive=True)
        if not pf_files:
            # If still no .pf files, try to find any text files that might contain IPs
            possible_files = glob.glob(os.path.join(temp_dir, "**", "snort*"), recursive=True)
            if possible_files:
                pf_files = possible_files
            else:
                raise FileNotFoundError("No snort_block.pf or similar file found after extraction")
        pf_file = pf_files[0]
    else:
        # If no .tar or .pf files, look for any file that might contain "snort" or "block"
        possible_files = glob.glob(os.path.join(temp_dir, "**", "*snort*"), recursive=True) + \
                         glob.glob(os.path.join(temp_dir, "**", "*block*"), recursive=True)
        
        if not possible_files:
            # List all extracted files for debugging
            all_files = []
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    all_files.append(os.path.join(root, file))
            
            print("All extracted files:")
            for file in all_files[:10]:  # Print first 10 files
                print(f"  - {file}")
                
            raise FileNotFoundError("No .tar, .pf, or snort files found after extraction")
        
        print(f"Found possible snort/block file: {possible_files[0]}")
        pf_file = possible_files[0]
    
    # Copy content to destination
    target_file = os.path.join(extract_dir, "snort.txt")
    with open(pf_file, 'r') as src, open(target_file, 'w') as dst:
        content = src.read()
        dst.write(content)
        print(f"File contains {len(content.splitlines())} lines")
    
    print(f"Successfully extracted to {target_file}")
    
    # Clean up temporary files
    shutil.rmtree(temp_dir)
    
    return target_file


def extract_ip_set_from_file(file_path):
    """Extract IP addresses from a file and return as a set"""
    ip_set = set()
    try:
        with open(file_path) as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                    
                try:
                    # Try to extract IP from line if it's not just a plain IP
                    ip_match = re.search(r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})', line)
                    if ip_match:
                        ip = ip_match.group(1)
                    else:
                        ip = line
                        
                    # Validate IP
                    ip = str(ipaddress.ip_address(ip))
                    ip_set.add(ip)
                except ValueError:
                    # Skip lines that don't contain valid IPs
                    continue
        
        print(f"Extracted {len(ip_set)} unique IP addresses from {file_path}")
        return ip_set
    except Exception as e:
        print(f"Error reading IP file {file_path}: {e}")
        return set()


def compare_ip_files(today_file, previous_file):
    """
    Compare two IP files and return the set of new IPs that appear in today's file
    but not in the previous file
    """
    # Handle case when there's no previous file for comparison
    if not previous_file:
        print("No previous file available for comparison. Processing all IPs in current file.")
        return extract_ip_set_from_file(today_file)
    
    # Extract IP sets from both files
    today_ips = extract_ip_set_from_file(today_file)
    previous_ips = extract_ip_set_from_file(previous_file)
    
    # Find IPs that are in today's file but not in previous file
    new_ips = today_ips - previous_ips
    
    print(f"Found {len(new_ips)} new IP addresses not present in previous file")
    print(f"Skipping {len(today_ips.intersection(previous_ips))} IP addresses that were already processed")
    
    return new_ips


def process_ip_addresses_from_set(ip_set, output_dir, master_xlsx_path, sheet_name):
    """Process a set of IP addresses and add them to an Excel workbook"""
    if not ip_set:
        print("No IP addresses to process.")
        return False
        
    print(f"Processing {len(ip_set)} IP addresses")
    
    # === Fetch Data ===
    handler = ipinfo.getHandler(IPINFO_ACCESS_TOKEN)
    data_list = []
    for ip in ip_set:
        try:
            details = handler.getDetails(ip, timeout=5)
            data = details.all
            filtered = {field: data.get(field, '') for field in IPINFO_FIELDS}
            data_list.append(filtered)
            print(f"Retrieved information for IP: {ip}")
        except Exception as e:
            print(f"Error fetching info for {ip}: {e}")
            # Add the IP with empty values for other fields
            filtered = {field: '' for field in IPINFO_FIELDS}
            filtered['ip'] = ip
            data_list.append(filtered)
        time.sleep(1)  # Respect API rate limits

    df = pd.DataFrame(data_list)
    
    # === Load existing workbook or create new one ===
    if not os.path.exists(master_xlsx_path):
        wb = Workbook()
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        wb.save(master_xlsx_path)
        print(f"Created new master file: {master_xlsx_path}")
    else:
        wb = load_workbook(master_xlsx_path)

    # Remove sheet if it already exists (optional)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)

    # === Write DataFrame to worksheet ===
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # === Auto-fit column widths ===
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2

    # === Highlight rows for selected organizations ===
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    org_col_index = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == 'org':
            org_col_index = idx
            break

    if org_col_index:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            org_cell = row[org_col_index - 1]
            if org_cell.value:
                for target_org in HIGHLIGHT_ORGS:
                    if target_org in org_cell.value:
                        for cell in row:
                            cell.fill = red_fill
                        break

    # === Add Excel Table Style with Unique Table Name ===
    if ws.max_row > 1:  # Only add table if there's data
        table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        table_name = f"IPData_{sheet_name.replace('-', '_').replace(':', '_')}"  # Safe table name
        table = Table(displayName=table_name, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium6",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        ws.add_table(table)

    # === Save workbook ===
    wb.save(master_xlsx_path)
    print(f"Data appended and formatted in: {master_xlsx_path}, sheet: {sheet_name}")
    return True


def process_ip_addresses(ip_file_path, output_dir, master_xlsx_path, sheet_name):
    """Process IP addresses from the extracted file and add them to an Excel workbook"""
    # Extract IP set from file
    ip_set = extract_ip_set_from_file(ip_file_path)
    
    if not ip_set:
        print("No valid IP addresses found in the file")
        return False
    
    # Process the IP set
    return process_ip_addresses_from_set(ip_set, output_dir, master_xlsx_path, sheet_name)


def main():
    """Main function to coordinate the entire workflow"""
    start_time = time.time()
    
    # Generate today's date for sheet name
    today = datetime.datetime.now().strftime('%d_%m_%Y')
    master_xlsx_path = os.path.join(OUTPUT_DIR, "master.xlsx")

    # Create output directory if it doesn't exist
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    print("=" * 60)
    print("STEP 1: DOWNLOADING SNORT BLOCKED HOSTS FILE")
    print("=" * 60)
    
    # Download the blocked hosts file
    downloaded_file = login_and_download_blocked_hosts(WEBSITE_CREDENTIALS)
    
    if not downloaded_file:
        print("Failed to download the file. Trying to locate the most recent download.")
        try:
            # Try to find the latest downloaded file instead
            downloaded_file = find_latest_download(DOWNLOAD_DIR)
        except FileNotFoundError as e:
            print(f"Error: {e}")
            print("Exiting script.")
            return
    
    # Try to find the previous download file for comparison
    previous_file = find_previous_download(DOWNLOAD_DIR, downloaded_file)
    
    print("\n" + "=" * 60)
    print("STEP 2: EXTRACTING SNORT BLOCKED HOSTS FILES")
    print("=" * 60)
    
    try:
        # Extract the current tar.gz file
        extracted_file = extract_tar_gz(downloaded_file, OUTPUT_DIR)
        
        # Extract the previous tar.gz file if it exists
        extracted_previous_file = None
        if previous_file:
            # Use a different output filename to avoid overwriting
            temp_extract_dir = os.path.join(OUTPUT_DIR, "temp_previous")
            os.makedirs(temp_extract_dir, exist_ok=True)
            extracted_previous_file = extract_tar_gz(previous_file, temp_extract_dir)
    except Exception as e:
        print(f"Error during file extraction: {e}")
        return
    
    print("\n" + "=" * 60)
    print("STEP 3: COMPARING IP ADDRESSES")
    print("=" * 60)
    
    # Compare IP addresses and get only new ones
    new_ips = compare_ip_files(extracted_file, extracted_previous_file)
    
    print("\n" + "=" * 60)
    print("STEP 4: PROCESSING NEW IP ADDRESSES")
    print("=" * 60)
    
    # Process only the new IP addresses and update Excel
    success = process_ip_addresses_from_set(new_ips, OUTPUT_DIR, master_xlsx_path, sheet_name=today)
    
    # Clean up temporary extraction directory for previous file if it exists
    if previous_file and os.path.exists(os.path.join(OUTPUT_DIR, "temp_previous")):
        shutil.rmtree(os.path.join(OUTPUT_DIR, "temp_previous"))
    
    print("\n" + "=" * 60)
    print("STEP 5: CLEANING UP SNORT FILES")
    print("=" * 60)
    
    # Clean up old snort files, keeping only the 2 most recent ones
    cleanup_old_snort_files(DOWNLOAD_DIR, keep_latest=2)
    
    # Delete the snort.txt file
    snort_txt_path = os.path.join(OUTPUT_DIR, "snort.txt")
    if os.path.exists(snort_txt_path):
        try:
            os.remove(snort_txt_path)
            print(f"Deleted file: {snort_txt_path}")
        except Exception as e:
            print(f"Error deleting file {snort_txt_path}: {e}")
    else:
        print(f"File not found: {snort_txt_path}")
    
    if success:
        print("\n" + "=" * 60)
        print("SCRIPT COMPLETED SUCCESSFULLY")
        print("=" * 60)
    else:
        print("\n" + "=" * 60)
        print("SCRIPT COMPLETED WITH ERRORS")
        print("=" * 60)
    
    elapsed_time = time.time() - start_time
    print(f"Total execution time: {elapsed_time:.2f} seconds")


if __name__ == "__main__":
    main()